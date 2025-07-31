from django.shortcuts import render, redirect, get_object_or_404
from django.core.mail import EmailMessage
from django.conf import settings
from .models import Company, Template, NoticeLog
from .forms import CompanyForm
from django.http import HttpResponse
import pandas as pd
from django.utils import timezone
from django.template.loader import render_to_string
import pdfkit
import openpyxl
from openpyxl.utils import get_column_letter
import os
from pathlib import Path
from collections import defaultdict
from datetime import date
from django.contrib import messages

def clean_cell(value):
    return None if pd.isna(value) or str(value).strip().lower() == "nan" else str(value).strip()

def generate_notice(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            return HttpResponse("Missing input: Please upload an Excel file.", status=400)

        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            return HttpResponse(f"Failed to read Excel file: {str(e)}", status=400)

        # Create date-based folder
        date_folder = timezone.now().strftime('%d-%m-%y')
        output_dir = Path(settings.MEDIA_ROOT) / "notices" / date_folder
        output_dir.mkdir(parents=True, exist_ok=True)

        for index, row in df.iterrows():
            try:
                # Extract required data
                company_name = row['Company'].strip()
                template_type = row['Template Type'].strip()
                company = Company.objects.get(name__iexact=company_name)
                template = Template.objects.get(company=company, template_type__iexact=template_type)

                context = {
                    'customer_name': row['Name'],
                    'customer_address': row['Address'],
                    'email': row['Email'],
                    'loan_account': row['Loan Account'],
                    'installments': row['Installments'],
                    'overdue_amount': row['Overdue Amount'],
                    'month_year': row['Month Year'],
                    'costs': row['Costs'],
                    'advocate_name': row['Advocate Name'],
                    'loan_date': row['Loan Date'],
                    'installment_amount': row['Installment Amount'],
                    'loan_amount': row['Loan Amount'],
                    'start_date': row['Start Date'],
                    'due_dates': row['Default Dates'],
                    'default_dates': row['Default Dates'],
                    'company_name': company.name,
                    'company_address': company.address,
                    'today_date': timezone.now().strftime('%d-%m-%Y'),
                    'due_date': timezone.now().strftime('%d-%m-%Y'),
                    'law_firm_name': clean_cell(row.get('Law Firm Name')),
                    'law_firm_address': clean_cell(row.get('Law Firm Address'))

                }

                # Generate formatted text
                notice_text = template.body.format(**context)
                email_body = template.email_body.format(**context)
                subject = template.subject.format(**context)

                # Render HTML for PDF
                html = render_to_string('notice_pdf.html', {
                    'subject': subject,
                    'notice_text': notice_text,
                    'law_firm_name': context['law_firm_name'],
                    'law_firm_address': context['law_firm_address'],

                })

                # Define PDF output path
                loan_id = row['Loan Account']
                pdf_filename = f"{loan_id}.pdf"
                pdf_path = output_dir / pdf_filename

                # Convert HTML to PDF using pdfkit
                pdfkit.from_string(html, str(pdf_path))

                # Email with PDF attachment
                email = EmailMessage(
                    subject=subject,
                    body=email_body,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[row['Email']],
                )
                email.attach_file(str(pdf_path))
                email.send()

                # Log success
                relative_pdf_path = os.path.relpath(pdf_path, settings.MEDIA_ROOT)
                NoticeLog.objects.create(
                    customer_name=row['Name'],
                    loan_account=row['Loan Account'],
                    template=template,
                    email=row['Email'],
                    company=company,
                    status='success',
                    pdf_file=relative_pdf_path,
                    timestamp=timezone.now()
                )

            except Exception as e:
                # Log failure
                NoticeLog.objects.create(
                    customer_name=row.get('Name', 'Unknown'),
                    loan_account=row.get('Loan Account', ''),
                    template=template if 'template' in locals() else None,
                    email=row.get('Email', ''),
                    company=company if 'company' in locals() else None,
                    status='failure',
                    reason=str(e),
                    timestamp=timezone.now()
                )
                print(f"[ERROR] Row {index+1}: {str(e)}")


        messages.success(request, "Notices have been generated. See Reports for more info.")
        return redirect('generate_notice')  #  This happens only on POST

    # For GET request: render the upload form
    return render(request, 'generate_notice.html')



def login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if username == 'Admin' and password == 'admin@789*':
            request.session['is_logged_in'] = True
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid username or password.")

    return render(request, 'login.html')

def logout(request):
    request.session.flush()
    return redirect('login')

def dashboard(request):
    if not request.session.get('is_logged_in'):
        return redirect('login')

    today = date.today()

    total_companies = Company.objects.count()
    total_templates = Template.objects.count()
    total_notices = NoticeLog.objects.count()

    successful_notices = NoticeLog.objects.filter(status='success').count()
    failed_notices = NoticeLog.objects.filter(status='failure').count()

    notices_success_today = NoticeLog.objects.filter(
        status='success',
        timestamp__date=today
    ).count()

    emails_sent_today = NoticeLog.objects.filter(
        status='success',
        timestamp__date=today
    ).exclude(email='').count()

    return render(request, 'dashboard.html', {
        'total_companies': total_companies,
        'total_templates': total_templates,
        'total_notices': total_notices,
        'successful_notices': successful_notices,
        'failed_notices': failed_notices,
        'notices_success_today': notices_success_today,
        'emails_sent_today': emails_sent_today,
    })



def manage_templates(request):
    templates = Template.objects.select_related('company').all()

    grouped = defaultdict(list)
    for template in templates:
        grouped[template.company].append(template)

    # Convert to list of tuples to make it template-friendly
    grouped_templates = list(grouped.items())

    return render(request, 'manage_templates.html', {
        'grouped_templates': grouped_templates
    })

def template_detail(request, id):
    template = get_object_or_404(Template, id=id)
    company = template.company
    return render(request, 'template_detail.html', {
        'template': template,
        'company': company,
        'email_body': template.email_body
    })


def manage_companies(request):
    companies = Company.objects.all()
    return render(request, 'manage_companies.html', {'companies': companies})


def add_company(request):
    if request.method == 'POST':
        form = CompanyForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('manage_companies')
    else:
        form = CompanyForm()
    return render(request, 'add_company.html', {'form': form})


def edit_company(request, id):
    company = get_object_or_404(Company, pk=id)
    if request.method == 'POST':
        form = CompanyForm(request.POST, instance=company)
        if form.is_valid():
            form.save()
            return redirect('manage_companies')
    else:
        form = CompanyForm(instance=company)
    return render(request, 'edit_company.html', {'form': form, 'company': company})


def delete_company(request, id):
    company = get_object_or_404(Company, pk=id)
    if request.method == 'POST':
        company.delete()
        return redirect('manage_companies')
    return render(request, 'delete_company.html', {'company': company})


def reports(request):
    success_count = NoticeLog.objects.filter(status='success').count()
    failure_count = NoticeLog.objects.filter(status='failure').count()

    # Group successful notices by generation date
    grouped_notices = defaultdict(list)
    success_logs = NoticeLog.objects.filter(status='success', pdf_file__isnull=False).order_by('-timestamp')

    for log in success_logs:
        date_str = log.timestamp.strftime('%d-%m-%Y')  # or use log.created_at if that field is used
        grouped_notices[date_str].append(log)

    return render(request, 'reports.html', {
        'success_count': success_count,
        'failure_count': failure_count,
        'grouped_notices': dict(grouped_notices),  # IMPORTANT: match this with the template key
    })


def download_failed_records_excel(request):
    from .models import NoticeLog

    # Fetch failed logs
    failed_logs = NoticeLog.objects.filter(status='failure').order_by('-timestamp')

    # Create an Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Notices"

    # Headers
    headers = [
        'Customer Name',
        'Email',
        'Loan Account',
        'Company',
        'Template',
        'Reason for Failure',
        'Created At',  # You can optionally rename this to 'Timestamp' to match the model
    ]
    ws.append(headers)

    # Data rows
    for log in failed_logs:
        row = [
            log.customer_name,
            log.email,
            log.loan_account,
            log.company.name if log.company else '',
            log.template.template_type if log.template else '',
            log.reason,
            log.timestamp.strftime('%d-%m-%Y %H:%M'),
        ]
        ws.append(row)

    # Auto column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # Create HTTP response with Excel content
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=failed_notices.xlsx'
    return response

def simplify_reason(reason):
    if not reason:
        return "Unknown Error"

    if isinstance(reason, dict):
        return "Invalid Email Address Format"

    if isinstance(reason, str):
        if "'float' object has no attribute 'strip'" in reason:
            return "A number was entered where text (like email) was expected"
        if "recipient address" in reason and "is not a valid" in reason:
            return "Invalid Email Address"
        if "Connection refused" in reason:
            return "Email Server Unreachable"
        if "timed out" in reason:
            return "Email Sending Timed Out"
        if "No template found" in reason:
            return "Missing Template"
        return reason[:150]

    return str(reason)[:150]
